import os
import random
import torch
import numpy as np
import time

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from torch.utils.data import DataLoader


def set_seed(args):
    random.seed(args.seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    torch.manual_seed(args.seed)
    np.random.seed(args.seed)
    return

def get_log_dir_datetime(args):
    all_runs = os.listdir(args.experiment_name)
    datetime =  time.strftime("%Y%m%d-%H%M%S")
    if args.name == None:
        exp_name = datetime
        log_dir = os.path.join(args.experiment_name, exp_name)
        args.name = exp_name
        args.paths.visualization = args.paths.visualization.format(exp_name, args.data_from)
        if args.train.wandb.enabled:
            args.train.wandb.name = exp_name
    elif args.name in all_runs:
        exp_name = args.name + "_" + datetime
        args.paths.visualization = args.paths.visualization.format(exp_name, args.data_from)
        log_dir = os.path.join(args.experiment_name, exp_name)
        args.name = exp_name
        if args.train.wandb.enabled:
            args.train.wandb.name = exp_name
    else:
        exp_name = args.name
        args.paths.visualization = args.paths.visualization.format(exp_name, args.data_from)
        log_dir = os.path.join(args.experiment_name, args.name)
    os.mkdir(log_dir)
    os.makedirs(args.paths.visualization)
    for action in args.visualizations.actions:
        os.mkdir(os.path.join(args.paths.visualization, action))
    return log_dir


def set_name(args):
    args.experiment_path = os.path.join(args.root_dir, args.dataset_name, args.group)
    if not os.path.exists(args.experiment_path):
        os.mkdir(args.experiment_path)
    all_runs = os.listdir(args.experiment_path)
    datetime =  time.strftime("%Y%m%d-%H%M%S")
    if args.name == None:
        args.name = datetime
    elif args.name in all_runs:
        args.name = args.name + "_" + datetime
    
    return args
    
def set_paths(args):
    args.experiment_path = os.path.join(args.experiment_path, args.name)
    args.config_path = os.path.join(args.experiment_path, 'config', 'config.yml')
    args.ckpt_path = os.path.join(args.experiment_path, 'ckpt', args.name + '_last.pt')
    args.best_path = os.path.join(args.experiment_path, 'ckpt', args.name + '_best.pt')
    if args.val_every != 0:
        args.epoch_path = os.path.join(args.experiment_path, 'ckpt', args.name + '_e{}.pt')
    args.viz_path = os.path.join(args.experiment_path, 'gifs')
    args.plot_path = os.path.join(args.experiment_path, 'plots')
    if not os.path.exists(args.experiment_path):
        os.mkdir(args.experiment_path)
    if not os.path.exists(os.path.join(args.experiment_path, 'config')):
        os.mkdir(os.path.join(args.experiment_path, 'config'))
    if not os.path.exists(os.path.join(args.experiment_path, 'ckpt')):
        os.mkdir(os.path.join(args.experiment_path, 'ckpt'))
    if not os.path.exists(args.plot_path):
        os.mkdir(args.plot_path)
    if not os.path.exists(args.viz_path):
        os.mkdir(args.viz_path)
        if args.dataset_name == 'h36m':
            for action in args.viz_actions:
                os.mkdir(os.path.join(args.viz_path, action))
    return args

def plot_losses(train_seq, val_seq, args):
    
    plt.figure()
    plt.plot(train_seq, 'r', label='Train loss')
    plt.savefig(os.path.join(args.plot_path, "train.png"))
    plt.close()
    
    plt.figure()
    plt.plot(val_seq, 'g', label='Val loss')
    plt.legend()
    plt.savefig(os.path.join(args.plot_path, "val.png"))
    plt.close()
    
    plt.figure()
    plt.plot(train_seq, 'r', label='Train loss')
    plt.plot(val_seq, 'g', label='Val loss')
    plt.legend()
    plt.savefig(os.path.join(args.plot_path, "train_and_val.png"))
    plt.close()
    
def line_prepender(filename, wandbURL):
    line = f'## See experiment at {wandbURL}\n\n'
    with open(filename, 'r+') as f:
        content = f.read()
        f.seek(0, 0)
        f.write(line.rstrip('\r\n') + '\n' + content)

def write_config(args):
    with open(args.config_path, 'w') as f:
        for k, v in args.__dict__.items():
            key = str(k)
            if type(v) == np.ndarray:
                val = str(v.tolist())
            elif v == None:
                val = '~'
            elif (v == True) or (v == False):
                val = str(v).lower()
            else:
                val = str(v)
            f.write(f"{key}: {val}\n")


def get_dataset_and_loader(args, split, return_dataset=False, actions=None, verbose=True):
    assert split in ['train', 'val', 'test', 0, 1, 2], 'split must be in ["train", "val", "test"]'
    if not args.global_translation:
        from utils import h36motion3d as datasets
    else:
        args.dim_used.sort()
        from utils import h36motion3dab as datasets
    if (split == 'train') or (split==0):
        split = 0
        bs = args.batch_size
        shuffle = True
    elif (split == 'val') or (split==1):
        split = 1
        bs = args.batch_size
        shuffle = False
    elif (split == 'test') or (split==2):
        split = 2
        bs = args.batch_size_test
        shuffle = False
    dataset = datasets.Datasets(args, actions=actions, split=split, verbose=verbose)
    if verbose:
        print('>>> Training dataset length: {:d}'.format(dataset.__len__()))
    data_loader = DataLoader(dataset, bs, shuffle=shuffle, num_workers=args.num_workers, pin_memory=True)
    
    if return_dataset:
        return dataset, data_loader
    return data_loader


def save_results(res_dict, args):
    if args.pred_loss:
        is_pred = args.coeff_pred
    else:
        is_pred = args.pred_loss
    if args.reco_loss:
        is_reco = args.coeff_reco
    else:
        is_reco = args.reco_loss
    if args.vel_loss:
        is_vel = args.coeff_vel
    else:
        is_vel = args.vel_loss
    if args.avo_loss:
        is_avo = args.coeff_avo
    else:
        is_avo = args.avo_loss
    data = [[args.name, args.n_epochs, args.batch_size, str(args.input_n) + " - " + str(args.output_n), args.n_pre,
              is_pred,
              is_reco,
              is_vel,
              is_avo,
              res_dict["walking"],
              res_dict["eating"],
              res_dict["smoking"],
              res_dict["discussion"],
              res_dict["directions"],
              res_dict["greeting"],
              res_dict["phoning"],
              res_dict["posing"],
              res_dict["purchases"],
              res_dict["sitting"],
              res_dict["sittingdown"],
              res_dict["takingphoto"],
              res_dict["waiting"],
              res_dict["walkingdog"],
              res_dict["walkingtogether"],
              res_dict["average"]]] # all values to report
    df = pd.DataFrame(data, columns=["name", "epochs", "batch size",  "in-out len", "dct", "L_pred coeff", "L_reco coeff", "L_vel coeff", "L_avo coeff", "walking", "eating", "smoking", "discussion", "directions",
                                     "greeting", "phoning", "posing", "purchases", "sitting", "sittingdown",
                                     "takingphoto", "waiting", "walkingdog", "walkingtogether", "average"]) # dataframe to store results
    
    sheet_name = args.dataset_name + '_' + args.metric
    if os.path.exists(args.table_path):
        FilePath = args.table_path
        ExcelWorkbook = load_workbook(FilePath)
        writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
        writer.book = ExcelWorkbook
        writer.sheets =  {ws.title: ws for ws in writer.book.worksheets}
        if sheet_name not in writer.sheets:
            ExcelWorkbook.create_sheet(title=sheet_name)
            writer.sheets =  {ws.title: ws for ws in writer.book.worksheets}
            startrow = writer.sheets[sheet_name].max_row
            df.to_excel(writer, sheet_name=sheet_name, index = False, header= True)
        else:
            startrow = writer.sheets[sheet_name].max_row
            df.to_excel(writer, sheet_name=sheet_name, startrow = startrow, index = False, header= False)
        # startr,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
        writer.save()
        writer.close()
    else:
        df.to_excel(args.table_path, sheet_name=sheet_name, index=False, header=True)